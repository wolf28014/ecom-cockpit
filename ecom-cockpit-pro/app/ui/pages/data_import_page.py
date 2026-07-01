"""
数据导入页 - Excel 拖拽导入
"""
from pathlib import Path
from datetime import date, datetime
from typing import List

from PySide6.QtWidgets import (
    QHBoxLayout, QVBoxLayout, QLabel, QPushButton, QComboBox, QTableWidget,
    QTableWidgetItem, QFileDialog, QMessageBox, QProgressBar, QGroupBox,
    QFormLayout, QHeaderView, QAbstractItemView, QFrame, QSizePolicy
)
from PySide6.QtCore import Qt, QMimeData, QUrl
from PySide6.QtGui import QDragEnterEvent, QDropEvent

from app.ui.pages.base_page import BasePage
from app.ui.components.section_card import SectionCard
from app.database.connection import get_session
from app.database.models import Store, DailyRecord
from app.config import PLATFORM_PROMOTION_FIELDS


class DropArea(QFrame):
    """拖拽区域"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("CardWidget")
        self.setAcceptDrops(True)
        self.setMinimumHeight(180)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        layout.setSpacing(8)

        icon = QLabel("📁")
        icon.setAlignment(Qt.AlignCenter)
        icon.setStyleSheet("font-size: 36px;")
        layout.addWidget(icon)

        self.title_label = QLabel("将 Excel 文件拖拽到此处")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("font-size: 16px; font-weight: 600; color: #1D1D1F;")
        layout.addWidget(self.title_label)

        self.desc_label = QLabel("支持 .xlsx / .xls 格式，最多 5MB")
        self.desc_label.setAlignment(Qt.AlignCenter)
        self.desc_label.setStyleSheet("font-size: 12px; color: #6E6E73;")
        layout.addWidget(self.desc_label)

        self.file_label = QLabel("")
        self.file_label.setAlignment(Qt.AlignCenter)
        self.file_label.setStyleSheet("font-size: 12px; color: #34C759; font-weight: 600;")
        layout.addWidget(self.file_label)

        self._file_path: Path = None
        self.on_file_dropped = None

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].toLocalFile().endswith((".xlsx", ".xls")):
                event.acceptProposedAction()
                self.setStyleSheet("QFrame#CardWidget { border: 2px dashed #0071E3; background: #F0F7FF; }")

    def dragLeaveEvent(self, event):
        self.setStyleSheet("")

    def dropEvent(self, event: QDropEvent):
        self.setStyleSheet("")
        file_path = event.mimeData().urls()[0].toLocalFile()
        self._file_path = Path(file_path)
        self.file_label.setText(f"已选择: {self._file_path.name}")
        if self.on_file_dropped:
            self.on_file_dropped(self._file_path)


class DataImportPage(BasePage):
    title = "Excel 自动导入"
    subtitle = "支持生意参谋/直通车/万相台/财务报表等模板"

    def __init__(self):
        super().__init__()

        # 顶部：店铺选择 + 模板选择
        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("目标店铺:"))
        self.store_combo = QComboBox()
        self.store_combo.setMinimumWidth(200)
        toolbar.addWidget(self.store_combo)

        toolbar.addSpacing(16)
        toolbar.addWidget(QLabel("模板类型:"))
        self.template_combo = QComboBox()
        self.template_combo.addItems([
            "通用模板（推荐）",
            "生意参谋报表",
            "直通车报表",
            "万相台报表",
            "财务报表",
        ])
        toolbar.addWidget(self.template_combo)

        toolbar.addStretch()
        self.add_layout(toolbar)

        # 拖拽区
        self.drop_area = DropArea()
        self.drop_area.on_file_dropped = self._on_file_dropped
        self.add_widget(self.drop_area)

        # 按钮区
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        browse_btn = QPushButton("浏览文件")
        browse_btn.clicked.connect(self._on_browse)
        btn_layout.addWidget(browse_btn)

        self.preview_btn = QPushButton("预览数据")
        self.preview_btn.clicked.connect(self._on_preview)
        btn_layout.addWidget(self.preview_btn)

        self.import_btn = QPushButton("开始导入")
        self.import_btn.setObjectName("PrimaryButton")
        self.import_btn.clicked.connect(self._on_import)
        btn_layout.addWidget(self.import_btn)

        self.add_layout(btn_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        self.add_widget(self.progress)

        # 预览表
        preview_section = SectionCard("数据预览", "导入前请确认数据无误")
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(8)
        self.preview_table.setHorizontalHeaderLabels([
            "日期", "销售额", "订单数", "退款金额", "退款订单", "推广费", "成本", "操作"
        ])
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        preview_section.add_widget(self.preview_table)
        self.add_widget(preview_section)

        # 模板下载
        download_section = SectionCard("模板下载", "首次使用请下载模板填写")
        download_layout = QHBoxLayout()

        info_label = QLabel("通用模板包含：日期、销售额、订单数、退款金额、退款订单数、推广费、商品成本、运费、包装、人工、房租、其他")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #6E6E73; font-size: 13px;")
        download_layout.addWidget(info_label)

        download_layout.addStretch()

        download_btn = QPushButton("下载通用模板")
        download_btn.setObjectName("PrimaryButton")
        download_btn.clicked.connect(self._on_download_template)
        download_layout.addWidget(download_btn)

        download_section.add_layout(download_layout)
        self.add_widget(download_section)

        self._preview_data: List[dict] = []
        self._load_stores()

    def _load_stores(self):
        self.store_combo.clear()
        with get_session() as session:
            stores = list(session.query(Store).filter(Store.is_active == True).order_by(Store.id).all())
            for s in stores:
                self.store_combo.addItem(f"{s.name}", s.id)

    def _on_browse(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            self.drop_area.file_label.setText(f"已选择: {Path(file_path).name}")
            self.drop_area._file_path = Path(file_path)

    def _on_file_dropped(self, file_path: Path):
        pass  # 已在 DropArea 中显示

    def _parse_excel(self, file_path: Path) -> List[dict]:
        """解析 Excel 文件为 dict 列表"""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        # 第一行作为表头
        headers = [str(h).strip() if h else "" for h in rows[0]]

        # 字段映射（模糊匹配）
        field_map = {
            "日期": ["日期", "date", "时间"],
            "销售额": ["销售额", "销售", "GMV", "sales"],
            "订单数": ["订单数", "订单", "orders"],
            "退款金额": ["退款金额", "退款"],
            "退款订单": ["退款订单数", "退款订单"],
            "推广费": ["推广费", "推广", "promotion"],
            "商品成本": ["商品成本", "成本"],
            "运费": ["运费"],
            "包装": ["包装"],
            "人工": ["人工"],
            "房租": ["房租"],
            "其他": ["其他"],
        }

        col_idx = {}
        for field, aliases in field_map.items():
            for idx, header in enumerate(headers):
                if any(alias in header for alias in aliases):
                    col_idx[field] = idx
                    break

        data = []
        for row in rows[1:]:
            if not row or not any(row):
                continue
            item = {}
            for field, idx in col_idx.items():
                val = row[idx] if idx < len(row) else None
                if field == "日期":
                    if isinstance(val, datetime):
                        item["日期"] = val.date()
                    elif isinstance(val, date):
                        item["日期"] = val
                    elif val:
                        try:
                            item["日期"] = datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
                        except Exception:
                            continue  # 跳过无法解析的行
                else:
                    try:
                        item[field] = float(val) if val else 0
                    except (ValueError, TypeError):
                        item[field] = 0
            if "日期" in item:
                data.append(item)
        return data

    def _on_preview(self):
        if not self.drop_area._file_path or not self.drop_area._file_path.exists():
            QMessageBox.warning(self, "提示", "请先选择 Excel 文件")
            return

        try:
            self._preview_data = self._parse_excel(self.drop_area._file_path)
            if not self._preview_data:
                QMessageBox.warning(self, "提示", "未解析到有效数据，请检查文件格式")
                return

            self.preview_table.setRowCount(len(self._preview_data))
            for row, item in enumerate(self._preview_data):
                self.preview_table.setItem(row, 0, QTableWidgetItem(str(item.get("日期", ""))))
                self.preview_table.setItem(row, 1, QTableWidgetItem(f"¥{item.get('销售额', 0):,.2f}"))
                self.preview_table.setItem(row, 2, QTableWidgetItem(str(int(item.get("订单数", 0)))))
                self.preview_table.setItem(row, 3, QTableWidgetItem(f"¥{item.get('退款金额', 0):,.2f}"))
                self.preview_table.setItem(row, 4, QTableWidgetItem(str(int(item.get("退款订单", 0)))))
                promo = sum(item.get(k, 0) for k in ["推广费"])
                cost = sum(item.get(k, 0) for k in ["商品成本", "运费", "包装", "人工", "房租", "其他"])
                self.preview_table.setItem(row, 5, QTableWidgetItem(f"¥{promo:,.2f}"))
                self.preview_table.setItem(row, 6, QTableWidgetItem(f"¥{cost:,.2f}"))
                self.preview_table.setItem(row, 7, QTableWidgetItem("待导入"))

            QMessageBox.information(self, "成功", f"已解析 {len(self._preview_data)} 条数据，请确认后点击导入")
        except Exception as e:
            QMessageBox.critical(self, "失败", f"解析失败: {e}")

    def _on_import(self):
        if not self._preview_data:
            QMessageBox.warning(self, "提示", "请先预览数据")
            return

        store_id = self.store_combo.currentData()
        if not store_id:
            QMessageBox.warning(self, "提示", "请选择目标店铺")
            return

        self.progress.setVisible(True)
        self.progress.setRange(0, len(self._preview_data))

        try:
            with get_session() as session:
                store = session.get(Store, store_id)
                promo_fields = PLATFORM_PROMOTION_FIELDS.get(store.platform, ["其他"])

                for i, item in enumerate(self._preview_data):
                    sales = item.get("销售额", 0)
                    orders = int(item.get("订单数", 0))
                    refund = item.get("退款金额", 0)
                    refund_orders = int(item.get("退款订单", 0))
                    promo_total = item.get("推广费", 0)

                    goods_cost = item.get("商品成本", 0)
                    shipping = item.get("运费", 0)
                    package = item.get("包装", 0)
                    labor = item.get("人工", 0)
                    rent = item.get("房租", 0)
                    other = item.get("其他", 0)

                    cost_data = {
                        "商品成本": goods_cost, "运费": shipping, "包装": package,
                        "人工": labor, "房租": rent, "其他": other,
                    }
                    cost_total = sum(cost_data.values())

                    # 推广预算分配到第一个字段
                    promo_data = {f: 0 for f in promo_fields}
                    if promo_fields:
                        promo_data[promo_fields[0]] = promo_total

                    gross = sales - goods_cost - refund
                    net = gross - promo_total - shipping - package - labor - rent - other

                    # 检查是否已存在
                    existing = session.query(DailyRecord).filter(
                        DailyRecord.store_id == store_id,
                        DailyRecord.record_date == item["日期"],
                    ).first()

                    if existing:
                        record = existing
                    else:
                        record = DailyRecord(store_id=store_id, record_date=item["日期"])
                        session.add(record)

                    record.sales_amount = sales
                    record.order_count = orders
                    record.refund_amount = refund
                    record.refund_order_count = refund_orders
                    record.promotion_data = promo_data
                    record.promotion_total = promo_total
                    record.cost_data = cost_data
                    record.cost_total = cost_total
                    record.gross_profit = round(gross, 2)
                    record.net_profit = round(net, 2)
                    record.profit_rate = round(net / sales, 4) if sales else 0
                    record.roi = round(sales / promo_total, 2) if promo_total else 0
                    record.avg_order_value = round(sales / orders, 2) if orders else 0
                    record.profit_per_order = round(net / orders, 2) if orders else 0
                    record.refund_rate = round(refund_orders / orders, 4) if orders else 0
                    record.promotion_rate = round(promo_total / sales, 4) if sales else 0

                    self.progress.setValue(i + 1)

                session.commit()

            QMessageBox.information(self, "成功", f"成功导入 {len(self._preview_data)} 条数据")
            self.progress.setVisible(False)
        except Exception as e:
            QMessageBox.critical(self, "失败", f"导入失败: {e}")
            self.progress.setVisible(False)

    def _on_download_template(self):
        """下载通用模板"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存模板", "电商数据导入模板.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "经营数据"

        headers = ["日期", "销售额", "订单数", "退款金额", "退款订单数", "推广费",
                   "商品成本", "运费", "包装", "人工", "房租", "其他"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0071E3")
            cell.alignment = Alignment(horizontal="center")

        # 示例行
        today = date.today()
        sample = [today.isoformat(), 8500, 60, 400, 3, 1500, 3500, 180, 60, 300, 200, 100]
        for col, v in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=v)

        # 列宽
        for col in range(1, 13):
            ws.column_dimensions[chr(64 + col)].width = 14

        wb.save(file_path)
        QMessageBox.information(self, "成功", f"模板已保存到:\n{file_path}")

    def refresh(self):
        self._load_stores()
